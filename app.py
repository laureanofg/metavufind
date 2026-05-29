# -*- coding: utf-8 -*-
import asyncio, logging, io, os, re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
from urllib.parse import urljoin

import httpx, yaml
from bs4 import BeautifulSoup
from fastapi import FastAPI, Form, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("app")
BASE_DIR = Path(__file__).parent

CONFIG_YAML = """
base_url: "https://descubridor.americana.edu.co"
max_results_per_target: 50
results_per_page: 20
targets:
  - id: ebsco_academic_search_complete
    name: "EBSCO Academic Search Complete"
    description: "Base de datos academica multidisciplinaria de EBSCO"
    url_template: "/Ebscoacademicsearchcomplete/Search?lookfor={query}&page={page}"
    selectors:
      container: "li.result"
      title: "a.title.getFull"
      author: "a.result-author"
      record_url: "a.hidden.full-record-link.icon-link"
      fulltext_url: "a.fulltext.icon-link"
      format: "div.result-formats span.format"
  - id: ebsco_ebook_academic_collection
    name: "EBSCO eBook Academic Collection"
    description: "Coleccion de libros electronicos academicos de EBSCO"
    url_template: "/Ebscoebookacademiccollection/Search?lookfor={query}&page={page}"
    selectors:
      container: "li.result"
      title: "a.title.getFull"
      author: "a.result-author"
      record_url: "a.hidden.full-record-link.icon-link"
      fulltext_url: "a.fulltext.icon-link"
      format: "div.result-formats span.format"
  - id: legisxperta
    name: "LegisXperta"
    description: "Plataforma de informacion juridica de Legis"
    url_template: "/Legisxperta/Search?lookfor={query}&page={page}"
    selectors:
      container: "li.result"
      title: "a.title.getFull"
      author: "a.result-author"
      record_url: "a.hidden.full-record-link.icon-link"
      fulltext_url: "a.fulltext.icon-link"
      format: "div.result-formats span.format"
  - id: metarevistas
    name: "MetaRevistas"
    description: "Metabuscador de revistas academicas"
    url_template: "/Metarevistas/Search?lookfor={query}&page={page}"
    selectors:
      container: "li.result"
      title: "a.title.getFull"
      author: "a.result-author"
      record_url: "a.hidden.full-record-link.icon-link"
      fulltext_url: "a.fulltext.icon-link"
      format: "div.result-formats span.format"
  - id: digitalia
    name: "Digitalia"
    description: "Plataforma de libros electronicos en espanol"
    url_template: "/Search/Results?filter%5B%5D=collection%3A%22Digitalia%22&lookfor={query}&page={page}"
    selectors:
      container: "li.result"
      title: "a.title.getFull"
      author: "a.result-author"
      record_url: "a.hidden.full-record-link.icon-link"
      fulltext_url: "a.fulltext.icon-link"
      format: "div.result-formats span.format"
"""

@dataclass
class ScrapedRecord:
    title: str = ""; authors: str = ""; record_url: str = ""; fulltext_url: str = ""; format_type: str = ""

@dataclass
class TargetResult:
    target_id: str; target_name: str
    records: list = field(default_factory=list)
    total_count: int = 0; error: Optional[str] = None

@dataclass
class TargetConfig:
    id: str; name: str; description: str; url_template: str; selectors: dict

@dataclass
class AppConfig:
    base_url: str; max_results_per_target: int; results_per_page: int; targets: list

def load_config() -> AppConfig:
    raw = yaml.safe_load(CONFIG_YAML)
    targets = [TargetConfig(id=t["id"], name=t["name"], description=t.get("description",""), url_template=t["url_template"], selectors=t["selectors"]) for t in raw["targets"]]
    return AppConfig(base_url=raw["base_url"], max_results_per_target=raw.get("max_results_per_target",50), results_per_page=raw.get("results_per_page",20), targets=targets)

config = load_config()

HEADERS = ["Titulo", "Autor(es)", "URL Registro", "URL Fulltext", "Tipo/Formato"]
COL_WIDTHS = [70, 50, 60, 60, 30]

def _sanitize_sheet_name(name: str) -> str:
    return re.sub(r"[\\\\\/*\[\]:?]", "", name)[:31]

def generate_excel(query: str, results: list) -> io.BytesIO:
    wb = Workbook(); wb.remove(wb.active)
    hf = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(vertical="top", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for result in results:
        ws = wb.create_sheet(title=_sanitize_sheet_name(result.target_name))
        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(row=1, column=ci, value=h); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri, rec in enumerate(result.records, 2):
            for ci, v in enumerate([rec.title, rec.authors, rec.record_url, rec.fulltext_url, rec.format_type], 1):
                c = ws.cell(row=ri, column=ci, value=v); c.alignment = ca; c.border = tb
        ws.auto_filter.ref = f"A1:E{len(result.records)+1}"; ws.freeze_panes = "A2"
    output = io.BytesIO(); wb.save(output); output.seek(0); return output

def _parse_record(record_html, selectors: dict, base_url: str) -> ScrapedRecord | None:
    try:
        soup = record_html if isinstance(record_html, BeautifulSoup) else BeautifulSoup(str(record_html), "html.parser")
        title_el = soup.select_one(selectors["title"]); title = title_el.get_text(strip=True) if title_el else ""
        author_els = soup.select(selectors["author"]); authors = "; ".join(a.get_text(strip=True) for a in author_els) if author_els else ""
        record_url_el = soup.select_one(selectors["record_url"]); record_url = urljoin(base_url, record_url_el.get("href","")) if record_url_el else ""
        fulltext_el = soup.select_one(selectors["fulltext_url"]); fulltext_url = fulltext_el.get("href","") if fulltext_el else ""
        format_els = soup.select(selectors["format"]); format_type = "; ".join(f.get_text(strip=True) for f in format_els) if format_els else ""
        if not title: return None
        return ScrapedRecord(title=title, authors=authors, record_url=record_url, fulltext_url=fulltext_url, format_type=format_type)
    except Exception:
        logger.exception("Error parsing record"); return None

async def _fetch_page(client: httpx.AsyncClient, url: str, timeout: int = 30) -> str | None:
    try:
        headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36","Accept":"text/html,application/xhtml+xml","Accept-Language":"es-ES,es;q=0.9"}
        resp = await client.get(url, headers=headers, timeout=timeout, follow_redirects=True)
        resp.raise_for_status(); return resp.text
    except httpx.HTTPStatusError as e:
        logger.error("HTTP error %s for %s", e.response.status_code, url); return None
    except Exception:
        logger.exception("Error fetching %s", url); return None

async def scrape_target(target: TargetConfig, query: str, timeout: int = 30) -> TargetResult:
    result = TargetResult(target_id=target.id, target_name=target.name)
    pages_needed = (config.max_results_per_target + config.results_per_page - 1) // config.results_per_page
    async with httpx.AsyncClient() as client:
        for page in range(1, pages_needed + 1):
            url = urljoin(config.base_url, target.url_template.format(query=query, page=page))
            logger.info("Scraping %s page %s", target.name, page)
            html = await _fetch_page(client, url, timeout)
            if html is None:
                if page == 1: result.error = f"No se pudo conectar a {target.name}"
                break
            soup = BeautifulSoup(html, "html.parser")
            record_elements = soup.select(target.selectors["container"])
            if not record_elements:
                if page == 1: result.error = f"No se encontraron resultados en {target.name}"
                break
            for record_el in record_elements:
                if len(result.records) >= config.max_results_per_target: break
                parsed = _parse_record(record_el, target.selectors, config.base_url)
                if parsed: result.records.append(parsed)
            if len(record_elements) < config.results_per_page: break
            await asyncio.sleep(0.3)
    return result

async def scrape_all_targets(query: str, timeout: int = 30) -> list:
    tasks = [scrape_target(target, query, timeout) for target in config.targets]
    return list(await asyncio.gather(*tasks))

app = FastAPI(title="MetaVufindScraping")

BASE_HTML = """<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>MetaVufindScraping</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/htmx.org@1.9.12/dist/htmx.min.js"></script>
<style>
.htmx-indicator{display:none}.htmx-request .htmx-indicator{display:inline}.htmx-request .htmx-indicator-hide{display:none}
.nav-tabs .nav-link{font-size:.85rem;white-space:nowrap}.table td{vertical-align:middle;font-size:.875rem}
.table-responsive{max-height:70vh;overflow-y:auto}.badge{font-size:.75rem;white-space:normal;text-align:left}
footer a{color:inherit}#results-area{min-height:100px}.tab-pane{min-height:200px}
</style></head><body>
<nav class="navbar navbar-expand-lg navbar-dark bg-primary mb-4"><div class="container">
<a class="navbar-brand fw-bold" href="/"><i class="bi bi-search"></i> MetaVufindScraping</a>
<span class="navbar-text text-white-50 small">Metabuscador academico via VUFIND</span></div></nav>
<div class="container">{content}</div>
<footer class="mt-5 py-3 text-center text-muted small border-top">
MetaVufindScraping - Busqueda en <a href="https://descubridor.americana.edu.co" target="_blank">Americana VUFIND</a></footer></body></html>"""

INDEX_HTML = """<div class="row justify-content-center"><div class="col-lg-10">
<div class="text-center mb-3"><h1 class="display-6 fw-bold text-primary">MetaVufindScraping</h1>
<p class="text-muted mb-2">Metabuscador academico que consulta multiples fuentes del sistema VUFIND.</p></div>
<div class="card shadow-sm mb-3"><div class="card-body p-3">
<form hx-post="/api/search" hx-target="#results-area" hx-indicator="#search-spinner" hx-disabled-elt="this">
<div class="row g-2 align-items-end">
<div class="col-md-7"><label class="form-label fw-semibold mb-1">Termino de busqueda</label>
<input type="text" class="form-control" name="query" placeholder="Ej: Colombia, inteligencia artificial..." value="{query}" required autofocus></div>
<div class="col-md-3"><label class="form-label fw-semibold mb-1">Max resultados</label>
<select class="form-select" name="max_results">
<option value="20" selected>20</option><option value="50">50</option><option value="100">100</option></select></div>
<div class="col-md-2"><button type="submit" class="btn btn-primary w-100">
<span id="search-spinner" class="htmx-indicator"><span class="spinner-border spinner-border-sm"></span></span>
<span class="htmx-indicator-hide"><i class="bi bi-search"></i> Buscar</span></button></div></div></form></div></div>
<details class="mb-2"><summary class="text-muted small cursor-pointer"><i class="bi bi-database"></i> Fuentes consultadas ({targets_count})</summary>
<div class="row g-1 mt-1">{targets_html}</div></details>
<div id="results-area"></div></div></div>"""

TARGETS_SOURCE = """<div class="col-md-4 col-6"><small class="text-muted"><i class="bi bi-globe2 text-primary me-1"></i>{name}</small></div>"""

RESULTS_CONTAINER = """<div id="results-container">
<div class="d-flex justify-content-between align-items-center mb-2">
<h5 class="mb-0"><i class="bi bi-list-check"></i> Resultados: <span class="text-primary">"{query}"</span></h5>
<a href="/api/export?query={query}&max_results={max_results}" class="btn btn-sm btn-success" download><i class="bi bi-file-earmark-excel"></i> Excel</a></div>
<ul class="nav nav-tabs mb-0" id="resultsTabs">{tabs}</ul><div class="tab-content">{tab_content}</div>
<script>
(function(){
    var btns = document.querySelectorAll('#resultsTabs .nav-link');
    btns.forEach(function(btn){ htmx.trigger(btn, 'load-tab'); });
})();
</script></div>"""

TAB_HEADER = """<li class="nav-item"><button class="nav-link {active}" id="tab-{tid}" data-bs-toggle="tab"
data-bs-target="#content-{tid}" type="button" role="tab"
hx-get="/api/search/{tid}?query={query}&max_results={mr}"
hx-trigger="load-tab once, click once" hx-target="#content-{tid}" hx-indicator="#indicator-{tid}"
hx-swap="innerHTML">{name}</button></li>"""
TAB_PANE = """<div class="tab-pane fade {active}" id="content-{tid}" role="tabpanel">
<div id="indicator-{tid}" class="text-center py-4"><div class="spinner-border spinner-border-sm text-primary"></div>
<small class="d-block mt-1 text-muted">Cargando {name}...</small></div></div>"""

TAB_RESULTS = """<div>{content}</div>"""

def make_tab_table(records, error, target_name):
    if error: return f'<div class="alert alert-danger"><i class="bi bi-exclamation-triangle"></i> {error}</div>'
    if not records: return f'<div class="alert alert-warning"><i class="bi bi-info-circle"></i> No se encontraron resultados en {target_name}.</div>'
    rows = ""
    for i, rec in enumerate(records):
        ft = f'<span class="badge bg-secondary">{rec.format_type}</span>' if rec.format_type else '<span class="text-muted">-</span>'
        rl = f'<a href="{rec.record_url}" target="_blank" class="btn btn-outline-secondary btn-sm"><i class="bi bi-box-arrow-up-right"></i> Ver</a>' if rec.record_url else '<span class="text-muted">-</span>'
        fl = f'<a href="{rec.fulltext_url}" target="_blank" class="btn btn-outline-success btn-sm"><i class="bi bi-file-text"></i> Fulltext</a>' if rec.fulltext_url else '<span class="text-muted">-</span>'
        rows += f'<tr><td class="text-center text-muted small">{i+1}</td><td><span class="fw-semibold">{rec.title}</span></td><td><span class="small">{rec.authors or "-"}</span></td><td>{ft}</td><td>{rl}</td><td>{fl}</td></tr>'
    return f'<div class="d-flex justify-content-between align-items-center mb-2"><small class="text-muted">{len(records)} resultados</small></div><div class="table-responsive"><table class="table table-striped table-hover table-sm align-middle"><thead class="table-dark"><tr><th>#</th><th>Titulo</th><th>Autor(es)</th><th>Tipo</th><th>Registro</th><th>Fulltext</th></tr></thead><tbody>{rows}</tbody></table></div>'

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, query: str = ""):
    targets_html = "".join(TARGETS_SOURCE.replace('{name}', t.name).replace('{desc}', t.description) for t in config.targets)
    return HTMLResponse(BASE_HTML.replace('{content}', INDEX_HTML.replace('{query}', query).replace('{targets_html}', targets_html).replace('{targets_count}', str(len(config.targets)))))

@app.post("/api/search", response_class=HTMLResponse)
async def search_results(request: Request, query: str = Form(..., min_length=1), max_results: int = Form(20, ge=10, le=200)):
    config.max_results_per_target = max_results
    tabs = "".join(TAB_HEADER.replace('{tid}', t.id).replace('{name}', t.name).replace('{active}', 'active' if i==0 else '') for i, t in enumerate(config.targets))
    tab_content = "".join(TAB_PANE.replace('{tid}', t.id).replace('{name}', t.name).replace('{mr}', str(max_results)).replace('{query}', query).replace('{active}', 'show active' if i==0 else '') for i, t in enumerate(config.targets))
    return HTMLResponse(RESULTS_CONTAINER.replace('{query}', query).replace('{max_results}', str(max_results)).replace('{tabs}', tabs).replace('{tab_content}', tab_content))

@app.get("/api/search/{target_id}", response_class=HTMLResponse)
async def search_target(request: Request, target_id: str, query: str = Query(..., min_length=1), max_results: int = Query(50, ge=10, le=200)):
    target = next((t for t in config.targets if t.id == target_id), None)
    if target is None: return HTMLResponse('<div class="alert alert-danger">Target no encontrado</div>', status_code=404)
    config.max_results_per_target = max_results
    try:
        result = await scrape_target(target, query)
    except Exception as e:
        return HTMLResponse(TAB_RESULTS.format(content=make_tab_table([], str(e), target.name)))
    return HTMLResponse(TAB_RESULTS.format(content=make_tab_table(result.records, result.error, target.name)))

@app.get("/api/export")
async def export_excel(query: str = Query(..., min_length=1), max_results: int = Query(50, ge=10, le=200)):
    config.max_results_per_target = max_results
    results = await scrape_all_targets(query)
    excel_bytes = generate_excel(query, results)
    filename = f"metavufind_{query.replace(' ', '_')}.xlsx"
    return StreamingResponse(excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
