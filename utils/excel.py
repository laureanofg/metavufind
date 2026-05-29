import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scrapers.engine import ScrapedRecord, TargetResult

HEADERS = ["Titulo", "Autor(es)", "URL Registro", "URL Fulltext", "Tipo/Formato"]
COL_WIDTHS = [70, 50, 60, 60, 30]
MAX_SHEET_NAME_LEN = 31


def _sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\\\/*\[\]:?]", "", name)
    return name[:MAX_SHEET_NAME_LEN]


def generate_excel(query: str, results: list[TargetResult]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for result in results:
        sheet_name = _sanitize_sheet_name(result.target_name)
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, record in enumerate(result.records, 2):
            values = [
                record.title,
                record.authors,
                record.record_url,
                record.fulltext_url,
                record.format_type,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        ws.auto_filter.ref = f"A1:E{len(result.records) + 1}"
        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
